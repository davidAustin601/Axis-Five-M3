from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import random
import csv



#DEFINITIONS
#Excel workbook and worksheet
workbook = Workbook()
worksheet = workbook.active
#file = "/Users/davidaustin/Desktop/Testing.xlsx"
#tempCSVFilePath = "/Users/davidaustin/Documents//Data/MARP/MARPMemberList.csv"
#tempMemberVacations = ["Bart Smith", "William Miller", "Cliff Hendley"]
globExcelOrange = "FFC000"

def readCSVFile(csvFilePath):
    csvData = []

    #read the csv file
    with open(csvFilePath, 'r') as file:
        reader = csv.reader(file)
        for row in reader:

            #get name from the row
            tempName = row[0]
        
            #make sure it's not the header
            if (tempName != "Name"):
                
                #get number of screen and convert to int
                tempNumScreens = int(row[4])
                tempSpecialNotes = row[6]
                
                #add name and num screens to the array
                tempArray = [tempName, tempNumScreens, tempSpecialNotes]
                csvData.append(tempArray)

                #if num screens = 2 then add again
                if (tempNumScreens == 2):
                    csvData.append(tempArray)

    #Shuffle the array
    random.shuffle(csvData)

    return csvData

def createCallList(csvFilePath, memberVacations, outputFile):

    #load the csv file from filePath
    csvData = readCSVFile(csvFilePath)
    
    print(len(csvData))

    #index for handling data
    count = 0

    #cycle through the csv file data row by row
    for item in csvData:

        #REFERENCE FOR CSVDATA FILE
        #Name(0), Email(1), Cell Phone(2), Alt Phone(3), Tests Per Month(4), Special Notes(6)

        rowIndex = count + 1

        #find row index (header row + current look index)
        #create new row right before newRowIndex
        worksheet.insert_rows(rowIndex)
        #add data to new row using newrowindex for row and 'row' for item in list
        #name
        worksheet.cell(row=rowIndex, column=1).value = item[0]
        #screens per month
        worksheet.cell(row=rowIndex, column=2).value = item[1]
        #Columns below are left blank
        #drug screen date
        #drug screen result
        #contact date
        #observed
        #Special Notes
        worksheet.cell(row=rowIndex, column=7).value = item[2]

        #edit the background to orange if they are on vacation during this month (use array parameter passed)
        blnOnVacation = False
        for member in memberVacations:
            if (item[0] == member):
                blnOnVacation = True
                break

        #if current row is on vacation
        if blnOnVacation:

            #Cycle through columns and change them to orange
            columnTicker = 0
            for tempCell in range(6):
                columnTicker += 1
                worksheet.cell(row=rowIndex, column=columnTicker).fill = PatternFill(fill_type='solid', start_color=globExcelOrange, end_color=globExcelOrange)


        #add to count
        count += 1


    #add columns to the excel file
    worksheet.insert_rows(1)
    #Name
    worksheet.cell(row=1, column=1).value = "Name"
    #Screens Per Month
    worksheet.cell(row=1, column=2).value = "Screens Per Month"
    #Drug Screen Date
    worksheet.cell(row=1, column=3).value = "Drug Screen Date"
    #Drug Screen Result
    worksheet.cell(row=1, column=4).value = "Drug Screen Result"
    #Contact Date
    worksheet.cell(row=1, column=5).value = "Contact Date"
    #Observed
    worksheet.cell(row=1, column=6).value = "Marked Observed"
    #Special Notes
    worksheet.cell(row=1, column=7).value = "Special Notes"

    #make column headers bold
    # Enumerate the cells in the second row
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True)

    #save the excel file
    workbook.save(outputFile)




